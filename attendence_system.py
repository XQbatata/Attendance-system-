from imutils.video import VideoStream
import face_recognition
import imutils
import pickle
import time
import cv2 as cv
import pandas as pd
import openpyxl
from datetime import datetime
import os
from openpyxl import load_workbook

# Loading the characteristics of traine but we need them in (top, right, bottom, left) order but we need them in (top, right, bottom, left) orderd faces from encodings.pickle file model
encodingsP = "encodings.pickle"

# https://github.com/opencv/opencv/blob/master/data/haarcascades/haarcascade_frontalface_default.xml
cascade = "haarcascade_frontalface_default.xml"

# load the known faces and embeddings
print("loading encodings + face detector to OpenCV classifier...")
data = pickle.loads(open(encodingsP, "rb").read())
detector = cv.CascadeClassifier(cascade)

print("Starting Pi Camera stream...")

def recognize_faces(number):
    # convert frame from BGR to grayscale for face detection
    # Convert frame from BGR to RGB for face recognition
    gray = cv.cvtColor(frame, cv.COLOR_BGR2GRAY)
    rgb = cv.cvtColor(frame, cv.COLOR_BGR2RGB)

    # detect faces in the grayscale frame
    rects = detector.detectMultiScale(gray, scaleFactor=1.1,
                                      minNeighbors=5, minSize=(30, 30),
                                      flags=cv.CASCADE_SCALE_IMAGE)

    # OpenCV returns bounding box coordinates in (x, y, w, h) order
    # but we need them in (top, right, bottom, left) order
    boxes = [(y, x + w, y + h, x) for (x, y, w, h) in rects]

    # compute the facial embeddings for each face bounding box
    encodings = face_recognition.face_encodings(rgb, boxes)
    names = []

    # loop over the facial embeddings
    for encoding in encodings:
        # attempt to match each face in the frame to the trained faces
        matches = face_recognition.compare_faces(data["encodings"],
                                                 encoding)
        name = "Unknown"  # if face is not recognized, then print Unknown

        # check to see if we have predefined person
        if True in matches:
            # find the indexes of all matched faces then initialize a
            # dictionary to count the total number of times each face
            # was matched
            matchedIdxs = [i for (i, b) in enumerate(matches) if b]
            counts = {}

            # loop over the matched indexes and maintain a count for
            # each recognized face face
            for i in matchedIdxs:
                name = data["names"][i]
                counts[name] = counts.get(name, 0) + 1

            # determine the recognized face with the largest number
            # of votes (note: in the event of an unlikely tie Python
            # will select first entry in the dictionary)
            name = max(counts, key=counts.get)

            # If someone in your dataset is identified, print their name on the screen
            date = datetime.today()
            print(name, date)
            df = pd.DataFrame([[name, date]],
                            index=[number], columns=['a', 'b',])

            # Check if the file exists and read the existing data
            file_name = 'data.xlsx'
            if os.path.exists(file_name):
                book = load_workbook(file_name)
                writer = pd.ExcelWriter(file_name, engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                startrow = writer.sheets['Sheet1'].max_row

                # Write the new data to the file without overwriting the existing data
                df.to_excel(writer, index=False, header=False, startrow=startrow)
            else:
                df.to_excel(file_name, index=False, header=False)

            writer.save()
        # update the list of names
        names.append(name)

    # loop over the recognized faces
    for ((top, right, bottom, left), name) in zip(boxes, names):
        # draw the predicted face name on the image - color is in BGR
        cv.rectangle(frame, (left, top), (right, bottom),
                     (255, 0, 0), 2)
        y = top - 15 if top - 15 > 15 else top + 15
        cv.putText(frame, name, (left, y), cv.FONT_HERSHEY_SIMPLEX,
                   .8, (0, 255, 0), 2)

# This....
# vs = VideoStream(src=0).start()

# Or this
vs = VideoStream(usePiCamera=True).start()

time.sleep(2.0)

while True:
    i = 0
    # Get the frame from the video stream and resize it to speedup processing
    frame = vs.read()
    frame = imutils.resize(frame, width=500)
    k = cv.waitKey(1)

    if k % 256 == 32:
        i = i+1
        recognize_faces(i)
    
    cv.imshow("Facial Recognition Stream", frame)
    key = cv.waitKey(1) & 0xFF

    # Stores the pressed key by the user.

    # Check if ESC is pressed
    if k % 256 == 27:
        print("ESC is pressed... Cancel")
        break

 
cv.destroyAllWindows()
vs.stop()